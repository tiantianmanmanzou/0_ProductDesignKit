#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DOCX转Excel脚本
将DOCX文件中的标题层级结构转换为Excel表格
"""

import sys
from pathlib import Path
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def extract_headings_from_docx(docx_path):
    """
    从DOCX文件中提取标题层级结构
    
    Args:
        docx_path: DOCX文件路径
        
    Returns:
        list: 包含标题信息的列表,每个元素为 (level, text, content)
    """
    doc = Document(docx_path)
    headings = []
    current_heading = None
    content_buffer = []
    
    for para in doc.paragraphs:
        # 检查是否为标题
        if para.style.name.startswith('Heading'):
            # 保存上一个标题的内容
            if current_heading is not None:
                headings.append({
                    'level': current_heading['level'],
                    'text': current_heading['text'],
                    'content': '\n'.join(content_buffer).strip()
                })
                content_buffer = []
            
            # 提取标题级别 (Heading 1 -> 1, Heading 2 -> 2, etc.)
            try:
                level = int(para.style.name.replace('Heading ', '').strip())
            except ValueError:
                level = 1
            
            current_heading = {
                'level': level,
                'text': para.text.strip()
            }
        else:
            # 收集正文内容
            if para.text.strip():
                content_buffer.append(para.text.strip())
    
    # 保存最后一个标题的内容
    if current_heading is not None:
        headings.append({
            'level': current_heading['level'],
            'text': current_heading['text'],
            'content': '\n'.join(content_buffer).strip()
        })
    
    return headings


def create_excel_from_headings(headings, output_path, max_level=6):
    """
    根据标题层级创建Excel文件
    
    Args:
        headings: 标题列表
        output_path: 输出Excel文件路径
        max_level: 最大标题级别
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "文档结构"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    content_font = Font(name='微软雅黑', size=10)
    content_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # 创建表头
    headers = []
    for i in range(1, max_level + 1):
        headers.append(f'{i}级标题')
    headers.append('正文内容')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    for i in range(1, max_level + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.column_dimensions[get_column_letter(max_level + 1)].width = 50
    
    # 填充数据
    row_idx = 2
    level_stack = [''] * max_level  # 用于记录各级别的当前标题
    
    for heading in headings:
        level = heading['level']
        text = heading['text']
        content = heading['content']
        
        # 更新当前级别及其子级别
        if level <= max_level:
            level_stack[level - 1] = text
            # 清空更低级别的标题
            for i in range(level, max_level):
                level_stack[i] = ''
        
        # 写入一行数据
        for col_idx in range(1, max_level + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=level_stack[col_idx - 1])
            cell.font = content_font
            cell.alignment = content_alignment
            cell.border = thin_border
        
        # 写入正文内容
        content_cell = ws.cell(row=row_idx, column=max_level + 1, value=content)
        content_cell.font = content_font
        content_cell.alignment = content_alignment
        content_cell.border = thin_border
        
        row_idx += 1
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(output_path)
    print(f"✓ Excel文件已保存到: {output_path}")


def main():
    """主函数"""
    # 输入文件路径
    input_file = "/Users/zhangxy/GenAI/DocPilot/docs/opex项目/重要系统对外接口逾期稽核功能-需求说明书20251202.docx"
    
    # 检查输入文件是否存在
    input_path = Path(input_file)
    if not input_path.exists():
        print(f"✗ 错误: 文件不存在 - {input_file}")
        sys.exit(1)
    
    # 生成输出文件路径
    output_path = input_path.with_suffix('.xlsx')
    
    print(f"开始转换...")
    print(f"输入文件: {input_file}")
    print(f"输出文件: {output_path}")
    print()
    
    # 提取标题
    print("正在提取文档标题...")
    headings = extract_headings_from_docx(input_path)
    print(f"✓ 共提取 {len(headings)} 个标题")
    
    # 分析标题级别
    levels = set(h['level'] for h in headings)
    max_level = max(levels) if levels else 1
    print(f"✓ 检测到的标题级别: {sorted(levels)}")
    print(f"✓ 最大标题级别: {max_level}")
    print()
    
    # 创建Excel文件
    print("正在生成Excel文件...")
    create_excel_from_headings(headings, output_path, max_level=max_level)
    print()
    print("转换完成!")


if __name__ == '__main__':
    main()
